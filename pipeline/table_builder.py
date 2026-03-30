import json
import os
import logging
from config import EXTRACTED_TEXT_FOLDER
from typing import List, Dict, Set, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Список полей, которые должны быть видимы в Excel (остальные скрыты)
VISIBLE_FIELDS = {
    'number', 'title', 'status', 'start_price', 'end_date',
    'delivery_place', 'customer', 'publish_date'
}


def generate_excel_table(input_dir: str, output_xlsx_path: str, jsonl_filename: str = "all_tenders_data.jsonl") -> None:
    """
    Читает JSONL файл из input_dir и создаёт Excel-файл с pretty-таблицей.
    Все колонки, кроме перечисленных в VISIBLE_FIELDS, будут скрыты.
    Номер тендера становится гиперссылкой на URL.
    Поля deep_rag_answers и deep_rag_score исключены.
    """
    jsonl_path = os.path.join(input_dir, jsonl_filename)
    if not os.path.exists(jsonl_path):
        logger.error("JSONL файл не найден: %s", jsonl_path)
        return

    records = []
    all_keys = set()
    try:
        with open(jsonl_path, 'r', encoding='utf-8') as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()
                if not line:
                    continue
                try:
                    data = json.loads(line)
                    if isinstance(data, dict):
                        records.append(data)
                        all_keys.update(data.keys())
                except json.JSONDecodeError as e:
                    logger.warning("Ошибка парсинга JSON в строке %d: %s", line_num, e)
    except Exception as e:
        logger.exception("Ошибка чтения файла %s: %s", jsonl_path, e)
        return

    if not records:
        logger.warning("Нет записей для Excel")
        return

    # Определяем поля для вывода
    fieldnames = sorted(all_keys)
    
    # Убираем технические/неинформативные поля
    tech_fields = ['url', 'description', 'total_price', 'total price', 'total_sum', 'files', 'tender_folder',
                   'relevance_score', 'relevance_explanation', 'reasoning', 'confidence']
    for key in tech_fields:
        if key in fieldnames:
            fieldnames.remove(key)

    # Разделяем поля на основные и deep_rag
    main_fields = []
    deep_rag_fields = []
    
    for field in fieldnames:
        if field.startswith('deep_rag_'):
            deep_rag_fields.append(field)
        else:
            main_fields.append(field)

    # Приоритетные поля – в начало (среди основных)
    priority_fields = ['number', 'title', 'status', 'start_price', 'end_date', 'delivery_place',
                       'customer', 'publish_date', 'positions', 'predicted_category','interest_score', 'is_interesting', 'interest_reasoning']
    
    for field in reversed(priority_fields):
        if field in main_fields:
            main_fields.remove(field)
            main_fields.insert(0, field)

    # Объединяем: сначала основные поля, потом deep_rag поля в конце
    fieldnames = main_fields + deep_rag_fields

    # Удаляем нежелательные deep_rag поля
    for unwanted in ['deep_rag_answers', 'deep_rag_score']:
        if unwanted in fieldnames:
            fieldnames.remove(unwanted)

    # Создаём книгу и лист
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Тендеры"

    # Заголовки
    header_names = {
        'number': 'Номер тендера',
        'title': 'Название',
        'status': 'Статус',
        'start_price': 'Начальная цена',
        'end_date': 'Окончание',
        'delivery_place': 'Место поставки',
        'customer': 'Заказчик',
        'publish_date': 'Дата публикации',
        'positions': 'Позиции',
        'predicted_category': 'Категория (ИИ)',
        'deep_rag_decision': 'Deep RAG: Решение',
        'deep_rag_reasoning': 'Deep RAG: Пояснение',
        'is_interesting': 'Интересен?',
        'interest_reasoning': 'Пояснение интереса',
    }
    
    headers = [header_names.get(key, key.replace('_', ' ').title()) for key in fieldnames]
    ws.append(headers)

    # Стиль заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Заполнение данных
    for rec in records:
        row = []
        for key in fieldnames:
            value = rec.get(key, '')
            if key == 'number' and rec.get('url'):
                value = str(value)
            elif key == 'positions' and isinstance(value, list):
                lines = []
                for pos in value:
                    name = pos.get('name', '')
                    price = pos.get('price', '')
                    quantity = pos.get('quantity', '')
                    unit = pos.get('unit', '')
                    qty_str = quantity
                    if unit:
                        qty_str += ' ' + unit
                    if price:
                        line = f"{name} - {price} - {qty_str}"
                    else:
                        line = f"{name} - {qty_str}"
                    lines.append(line)
                value = '\n'.join(lines) if lines else ''
            elif key == 'is_interesting':
                value = "Да" if value else "Нет"
            elif isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False, indent=2)
            else:
                value = str(value) if value is not None else ''
            row.append(value)

        ws.append(row)

        # ГИПЕРССЫЛКА НА НОМЕР ТЕНДЕРА
        if 'number' in fieldnames:
            number_col_idx = fieldnames.index('number') + 1
            cell = ws.cell(row=ws.max_row, column=number_col_idx)
            url = rec.get('url', '')
            if url:
                cell.hyperlink = url
                cell.style = 'Hyperlink'

    # Автоподбор ширины колонок
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Скрыть колонки, не входящие в список VISIBLE_FIELDS
    for col_idx, key in enumerate(fieldnames, start=1):
        col_letter = get_column_letter(col_idx)
        if key not in VISIBLE_FIELDS:
            ws.column_dimensions[col_letter].hidden = True

    # Сохраняем
    os.makedirs(os.path.dirname(output_xlsx_path) or '.', exist_ok=True)
    wb.save(output_xlsx_path)
    logger.info("Excel-таблица создана: %s, записей: %d", output_xlsx_path, len(records))


def generate_html_table(input_dir: str, output_html_path: str, jsonl_filename: str = "all_tenders_data.jsonl") -> None:
    """
    Читает JSONL файл из input_dir и создаёт HTML-файл с pretty-таблицей, фильтрами
    и кнопкой экспорта в Excel (с гиперссылками на номер тендера).
    Поля deep_rag_answers и deep_rag_score исключены.
    """
    records = []
    all_keys: Set[str] = set()

    jsonl_path = os.path.join(input_dir, jsonl_filename)

    if not os.path.exists(jsonl_path):
        logger.error("JSONL файл не найден: %s", jsonl_path)
        return

    try:
        with open(jsonl_path, 'r', encoding='utf-8') as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()
                if not line:
                    continue
                    
                try:
                    data = json.loads(line)
                    if isinstance(data, dict):
                        records.append(data)
                        all_keys.update(data.keys())
                    else:
                        logger.warning("Строка %d содержит не объект, а %s. Пропускаем.", 
                                     line_num, type(data).__name__)
                except json.JSONDecodeError as e:
                    logger.error("Ошибка парсинга JSON в строке %d: %s", line_num, e)
                    
    except Exception as e:
        logger.exception("Ошибка чтения файла %s: %s", jsonl_path, e)
        return

    if not records:
        logger.warning("Не найдено записей в файле %s", jsonl_path)
        return

    logger.info("Загружено %d записей из %s", len(records), jsonl_path)

    # Определяем порядок колонок
    fieldnames = sorted(all_keys)
    
    # Убираем ненужные колонки
    tech_fields = ['url', 'description', 'total_price', 'total price', 'total_sum', 'files', 'tender_folder',
                   'relevance_score', 'relevance_explanation', 'reasoning', 'confidence']
    for key in tech_fields:
        if key in fieldnames:
            fieldnames.remove(key)
    
    if 'delivery_place' not in fieldnames:
        fieldnames.append('delivery_place')

    # Разделяем поля на основные и deep_rag
    main_fields = []
    deep_rag_fields = []
    
    for field in fieldnames:
        if field.startswith('deep_rag_'):
            deep_rag_fields.append(field)
        else:
            main_fields.append(field)

    # Приоритетные поля – в начало (среди основных)
    priority_fields = [
        'number',
        'title',
        'status',
        'start_price',
        'end_date',
        'delivery_place',
        'customer',
        'publish_date',
        'positions',
        'predicted_category',
        'is_interesting',
        'interest_score',
        'interest_reasoning'
    ]
    
    for field in reversed(priority_fields):
        if field in main_fields:
            main_fields.remove(field)
            main_fields.insert(0, field)

    # Объединяем: сначала основные поля, потом deep_rag поля в конце
    fieldnames = main_fields + deep_rag_fields

    # Удаляем нежелательные deep_rag поля
    for unwanted in ['deep_rag_answers', 'deep_rag_score']:
        if unwanted in fieldnames:
            fieldnames.remove(unwanted)

    # Генерация HTML
    html_content = _build_html_table(records, fieldnames)
    os.makedirs(os.path.dirname(output_html_path) or '.', exist_ok=True)
    with open(output_html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

    logger.info("HTML-таблица создана: %s, записей: %d", output_html_path, len(records))


def _build_html_table(records: List[Dict[str, Any]], fieldnames: List[str]) -> str:
    """
    Формирует HTML-страницу с таблицей, фильтрами, счётчиком и экспортом в Excel.
    При экспорте номера тендеров становятся гиперссылками с синим цветом и подчёркиванием.
    Поля deep_rag_answers и deep_rag_score уже исключены из fieldnames.
    """
    # Fields to exclude from HTML display (they are still in records for filtering)
    exclude_from_html = {'is_interesting'}
    html_fieldnames = [f for f in fieldnames if f not in exclude_from_html]

    # Собираем уникальные категории для фильтра
    categories = set()
    for rec in records:
        cat = rec.get('predicted_category', '')
        if cat:
            categories.add(cat)
    categories = sorted(categories)

    # Заголовки (для отображения)
    header_names = {
        'number': 'Номер тендера',
        'title': 'Название',
        'status': 'Статус',
        'start_price': 'Начальная цена',
        'end_date': 'Окончание',
        'delivery_place': 'Место поставки',
        'customer': 'Заказчик',
        'publish_date': 'Дата публикации',
        'positions': 'Позиции',
        'predicted_category': 'Категория (ИИ)',
        'deep_rag_decision': 'Deep RAG: Решение',
        'deep_rag_reasoning': 'Deep RAG: Пояснение',
        'interest_reasoning': 'Пояснение интереса',
        'interest_score': 'Оценка интереса',
    }
    headers_ru = [header_names.get(k, k.replace('_', ' ').title()) for k in html_fieldnames]

    # Формируем строки HTML
    rows_html = []
    for idx, rec in enumerate(records, start=1):
        is_interesting = rec.get('is_interesting', False)
        category = rec.get('predicted_category', '') or 'Не классифицировано'
        tr_attrs = f'data-interest="{str(is_interesting).lower()}" data-category="{category}"'
        
        cells = [f'<td style="text-align: center; font-weight: bold;">{idx}']
        for key in html_fieldnames:
            # ---- НОМЕР ТЕНДЕРА КАК ССЫЛКА ----
            if key == 'number':
                number_value = rec.get(key, '')
                number_str = str(number_value) if number_value is not None else ''
                number_escaped = (number_str.replace('&', '&amp;')
                                            .replace('<', '&lt;')
                                            .replace('>', '&gt;')
                                            .replace('"', '&quot;')
                                            .replace("'", '&#39;'))
                url = rec.get('url', '')
                if url:
                    url_str = str(url)
                    url_escaped = (url_str.replace('&', '&amp;')
                                          .replace('<', '&lt;')
                                          .replace('>', '&gt;')
                                          .replace('"', '&quot;')
                                          .replace("'", '&#39;'))
                    cell_content = f'<a href="{url_escaped}" target="_blank" title="{number_escaped}">{number_escaped}</a>'
                else:
                    cell_content = number_escaped
                cells.append(f'<td style="white-space: nowrap;">{cell_content}')

            # ---- НАЗВАНИЕ ТЕНДЕРА (ТЕКСТ) ----
            elif key == 'title':
                title_value = rec.get(key, '')
                title_str = str(title_value) if title_value is not None else ''
                if len(title_str) > 500:
                    title_str = title_str[:500] + '…'
                title_escaped = (title_str.replace('&', '&amp;')
                                          .replace('<', '&lt;')
                                          .replace('>', '&gt;')
                                          .replace('"', '&quot;')
                                          .replace("'", '&#39;'))
                cells.append(f'<td style="min-width: 200px;">{title_escaped}')

            # ---- ЗАКАЗЧИК ----
            elif key == 'customer':
                customer_value = rec.get(key, '')
                if isinstance(customer_value, (dict, list)):
                    customer_value = json.dumps(customer_value, ensure_ascii=False, indent=2)
                else:
                    customer_value = str(customer_value) if customer_value is not None else ''
                if len(customer_value) > 200:
                    customer_value = customer_value[:200] + '…'
                customer_escaped = (customer_value.replace('&', '&amp;')
                                                   .replace('<', '&lt;')
                                                   .replace('>', '&gt;')
                                                   .replace('"', '&quot;')
                                                   .replace("'", '&#39;'))
                cells.append(f'<td style="min-width: 150px;">{customer_escaped}')

            # ---- МЕСТО ПОСТАВКИ ----
            elif key == 'delivery_place':
                place_value = rec.get(key, '')
                if isinstance(place_value, (dict, list)):
                    place_value = json.dumps(place_value, ensure_ascii=False, indent=2)
                else:
                    place_value = str(place_value) if place_value is not None else ''
                if len(place_value) > 200:
                    place_value = place_value[:200] + '…'
                place_escaped = (place_value.replace('&', '&amp;')
                                           .replace('<', '&lt;')
                                           .replace('>', '&gt;')
                                           .replace('"', '&quot;')
                                           .replace("'", '&#39;'))
                cells.append(f'<td style="min-width: 150px;">{place_escaped}')

            # ---- ПОЗИЦИИ (СПИСОК) ----
            elif key == 'positions':
                positions_data = rec.get(key, [])
                if isinstance(positions_data, list):
                    lines = []
                    for pos in positions_data:
                        name = pos.get('name', '')
                        name_escaped = (name.replace('&', '&amp;')
                                            .replace('<', '&lt;')
                                            .replace('>', '&gt;')
                                            .replace('"', '&quot;')
                                            .replace("'", '&#39;'))
                        price = pos.get('price', '')
                        if price:
                            price_escaped = (price.replace('&', '&amp;')
                                                  .replace('<', '&lt;')
                                                  .replace('>', '&gt;')
                                                  .replace('"', '&quot;')
                                                  .replace("'", '&#39;'))
                        else:
                            price_escaped = ''
                        quantity = pos.get('quantity', '')
                        unit = pos.get('unit', '')
                        qty_str = quantity
                        if unit:
                            qty_str += ' ' + unit
                        qty_escaped = (qty_str.replace('&', '&amp;')
                                              .replace('<', '&lt;')
                                              .replace('>', '&gt;')
                                              .replace('"', '&quot;')
                                              .replace("'", '&#39;'))
                        if price_escaped:
                            line = f"{name_escaped} - {price_escaped} - {qty_escaped}"
                        else:
                            line = f"{name_escaped} - {qty_escaped}"
                        lines.append(line)
                    cell_content = '<br>'.join(lines) if lines else ''
                else:
                    cell_content = str(positions_data)
                if len(cell_content) > 500:
                    cell_content = cell_content[:500] + '…'
                cells.append(f'<td style="white-space: pre-wrap; width: 500px;">{cell_content}')

            # ---- НАЧАЛЬНАЯ ЦЕНА ----
            elif key == 'start_price':
                price_value = rec.get(key, '')
                price_str = str(price_value) if price_value is not None else ''
                price_escaped = (price_str.replace('&', '&amp;')
                                         .replace('<', '&lt;')
                                         .replace('>', '&gt;')
                                         .replace('"', '&quot;')
                                         .replace("'", '&#39;'))
                cells.append(f'<td style="white-space: nowrap;"><strong>{price_escaped}</strong>')

            # ---- СТАТУС ----
            elif key == 'status':
                status_value = rec.get(key, '')
                status_str = str(status_value) if status_value is not None else ''
                status_escaped = (status_str.replace('&', '&amp;')
                                           .replace('<', '&lt;')
                                           .replace('>', '&gt;')
                                           .replace('"', '&quot;')
                                           .replace("'", '&#39;'))
                color_class = ''
                if 'Приём заявок' in status_str or 'прием заявок' in status_str.lower():
                    color_class = 'status-active'
                elif 'Завершён' in status_str or 'завершен' in status_str.lower():
                    color_class = 'status-completed'
                elif 'Отменён' in status_str or 'отменен' in status_str.lower():
                    color_class = 'status-cancelled'
                cells.append(f'<td class="{color_class}">{status_escaped}')

            # ---- ДАТЫ ----
            elif key in ('end_date', 'publish_date'):
                date_value = rec.get(key, '')
                date_str = str(date_value) if date_value is not None else ''
                date_escaped = (date_str.replace('&', '&amp;')
                                       .replace('<', '&lt;')
                                       .replace('>', '&gt;')
                                       .replace('"', '&quot;')
                                       .replace("'", '&#39;'))
                cells.append(f'<td style="white-space: nowrap;"><span class="date">{date_escaped}</span>')

            # ---- КАТЕГОРИЯ (ИИ) ----
            elif key == 'predicted_category':
                cat = rec.get(key, '')
                cat_str = str(cat) if cat is not None else ''
                cat_escaped = (cat_str.replace('&', '&amp;')
                                      .replace('<', '&lt;')
                                      .replace('>', '&gt;')
                                      .replace('"', '&quot;')
                                      .replace("'", '&#39;'))
                cells.append(f'<td style="white-space: nowrap;"><strong>{cat_escaped}</strong>')

            # ---- DEEP RAG DECISION ----
            elif key == 'deep_rag_decision':
                decision = rec.get(key, '')
                decision_str = str(decision) if decision is not None else ''
                decision_escaped = (decision_str.replace('&', '&amp;')
                                                 .replace('<', '&lt;')
                                                 .replace('>', '&gt;')
                                                 .replace('"', '&quot;')
                                                 .replace("'", '&#39;'))
                cells.append(f'<td style="white-space: nowrap;"><strong>{decision_escaped}</strong>')

            # ---- DEEP RAG REASONING ----
            elif key == 'deep_rag_reasoning':
                reasoning = rec.get(key, '')
                if reasoning is None:
                    reasoning = ''
                reasoning_str = str(reasoning)
                reasoning_escaped = (reasoning_str.replace('&', '&amp;')
                                                    .replace('<', '&lt;')
                                                    .replace('>', '&gt;')
                                                    .replace('"', '&quot;')
                                                    .replace("'", '&#39;'))
                if not reasoning_escaped:
                    reasoning_escaped = '—'
                cells.append(f'<td style="white-space: pre-wrap; max-width: 500px;">{reasoning_escaped}')

            # ---- ПОЯСНЕНИЕ ИНТЕРЕСА ----
            elif key == 'interest_reasoning':
                reasoning = rec.get(key, '')
                if reasoning is None:
                    reasoning = ''
                reasoning_str = str(reasoning)
                reasoning_escaped = (reasoning_str.replace('&', '&amp;')
                                                  .replace('<', '&lt;')
                                                  .replace('>', '&gt;')
                                                  .replace('"', '&quot;')
                                                  .replace("'", '&#39;'))
                if not reasoning_escaped:
                    reasoning_escaped = '—'
                cells.append(f'<td style="white-space: pre-wrap; max-width: 400px;">{reasoning_escaped}')

            # ---- ВСЕ ОСТАЛЬНЫЕ ПОЛЯ ----
            else:
                value = rec.get(key, '')
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False, indent=2)
                else:
                    value = str(value) if value is not None else ''
                if len(value) > 200:
                    value = value[:200] + '…'
                value = (value.replace('&', '&amp;')
                              .replace('<', '&lt;')
                              .replace('>', '&gt;')
                              .replace('"', '&quot;')
                              .replace("'", '&#39;'))
                cells.append(f'<td style="white-space: pre-wrap;">{value}')

        rows_html.append(f'<tr {tr_attrs}>' + ''.join(cells))

    # Построение HTML-заголовка таблицы с onclick для сортировки
    headers_html = '<th style="text-align: center;">№</th>'
    for i, header_ru in enumerate(headers_ru):
        headers_html += f'<th onclick="sortTable({i+1})" title="{html_fieldnames[i]}">{header_ru}</th>'

    # Блок фильтров с кнопкой экспорта и счётчиком
    filter_buttons_html = f'''
    <div class="filters" style="margin-bottom: 20px; display: flex; flex-direction: column; gap: 15px;">
        <div style="display: flex; gap: 15px; align-items: center; flex-wrap: wrap;">
            <div>
                <span style="font-weight: bold;">Интерес:</span>
                <button id="filter-all" class="filter-btn active">Все</button>
                <button id="filter-interesting" class="filter-btn">Только интересные</button>
                <button id="filter-not-interesting" class="filter-btn">Только неинтересные</button>
            </div>
            <div>
                <span style="font-weight: bold;">Категория:</span>
                <select id="category-filter">
                    <option value="all">Все категории</option>
                    {''.join(f'<option value="{cat}">{cat}</option>' for cat in categories)}
                </select>
            </div>
            <div>
                <span style="font-weight: bold;">Показано:</span>
                <span id="record-count" style="font-weight: bold;">0</span> из {len(records)}
            </div>
            <div>
                <span style="font-weight: bold;">Колонки:</span>
                <button id="show-all-columns" class="filter-btn active">Все поля</button>
                <button id="show-main-columns" class="filter-btn">Основные поля</button>
            </div>
            <button id="reset-filters" class="filter-btn">Сбросить фильтры</button>
            <button id="export-excel" class="filter-btn">📊 Скачать Excel (с фильтрами)</button>
        </div>
        <div id="active-filters-info" style="font-size: 0.9em; color: #2c3e50; background: #ecf0f1; padding: 8px 12px; border-radius: 6px;">
            Фильтры не применены
        </div>
    </div>
    '''

    # JS для фильтрации, сортировки, счётчика и экспорта
    js_filter = f'''
    <script src="https://cdn.jsdelivr.net/npm/xlsx-js-style@1.2.0/dist/xlsx.bundle.js"></script>
    <script>
        const headers = {json.dumps(headers_ru, ensure_ascii=False)};
        const fieldKeys = {json.dumps(html_fieldnames, ensure_ascii=False)};
        const allRecords = {json.dumps(records, ensure_ascii=False)};

        let sortColumn = -1;
        let sortDirection = 1;
        let currentColumnsMode = 'all';

        // Индексы основных полей (пропускаем первый столбец с номером)
        const mainFieldKeys = ['number', 'title', 'status', 'start_price', 'end_date',
                               'delivery_place', 'customer', 'publish_date', 'positions',
                               'predicted_category', 'deep_rag_decision'];
        const mainColumnIndices = [];
        for (let i = 0; i < fieldKeys.length; i++) {{
            if (mainFieldKeys.includes(fieldKeys[i])) {{
                mainColumnIndices.push(i + 1);
            }}
        }}

        function parseRussianDate(dateStr) {{
            if (!dateStr) return new Date(0);
            let parts = dateStr.split(' ');
            let datePart = parts[0];
            let timePart = parts[1] || '00:00:00';
            let dmy = datePart.split('.');
            if (dmy.length === 3) {{
                let year = parseInt(dmy[2], 10);
                let month = parseInt(dmy[1], 10) - 1;
                let day = parseInt(dmy[0], 10);
                let time = timePart.split(':');
                let hour = parseInt(time[0], 10);
                let minute = parseInt(time[1], 10);
                let second = parseInt(time[2], 10);
                return new Date(year, month, day, hour, minute, second);
            }}
            return new Date(dateStr);
        }}

        function sortTable(colIndex) {{
            const table = document.getElementById('dataTable');
            const tbody = table.tBodies[0];
            const rows = Array.from(tbody.rows);
            if (sortColumn === colIndex) {{
                sortDirection = -sortDirection;
            }} else {{
                sortColumn = colIndex;
                sortDirection = 1;
            }}
            rows.sort((a, b) => {{
                let aVal, bVal;
                if (colIndex === 0) {{
                    aVal = parseInt(a.cells[0].innerText);
                    bVal = parseInt(b.cells[0].innerText);
                }} else {{
                    aVal = a.cells[colIndex].innerText.trim();
                    bVal = b.cells[colIndex].innerText.trim();
                    const key = table.rows[0].cells[colIndex].getAttribute('title');
                    if (key === 'start_price') {{
                        aVal = parseFloat(aVal.replace(/[^\\d.,-]/g, '').replace(',', '.'));
                        bVal = parseFloat(bVal.replace(/[^\\d.,-]/g, '').replace(',', '.'));
                        if (isNaN(aVal)) aVal = -Infinity;
                        if (isNaN(bVal)) bVal = -Infinity;
                    }} else if (key === 'end_date' || key === 'publish_date') {{
                        aVal = parseRussianDate(aVal);
                        bVal = parseRussianDate(bVal);
                    }} else {{
                        aVal = aVal.toLowerCase();
                        bVal = bVal.toLowerCase();
                    }}
                }}
                if (aVal < bVal) return -sortDirection;
                if (aVal > bVal) return sortDirection;
                return 0;
            }});
            rows.forEach(row => tbody.appendChild(row));
            for (let i = 0; i < rows.length; i++) {{
                rows[i].cells[0].innerText = (i + 1).toString();
            }}
        }}

        function showAllColumns() {{
            currentColumnsMode = 'all';
            const rows = document.querySelectorAll('#dataTable tbody tr');
            rows.forEach(row => {{
                const cells = row.cells;
                for (let i = 1; i < cells.length; i++) {{
                    cells[i].style.display = '';
                }}
            }});
            const headers = document.querySelectorAll('#dataTable thead th');
            for (let i = 1; i < headers.length; i++) {{
                headers[i].style.display = '';
            }}
        }}

        function showMainColumns() {{
            currentColumnsMode = 'main';
            const rows = document.querySelectorAll('#dataTable tbody tr');
            rows.forEach(row => {{
                const cells = row.cells;
                for (let i = 1; i < cells.length; i++) {{
                    cells[i].style.display = mainColumnIndices.includes(i) ? '' : 'none';
                }}
            }});
            const headers = document.querySelectorAll('#dataTable thead th');
            for (let i = 1; i < headers.length; i++) {{
                headers[i].style.display = mainColumnIndices.includes(i) ? '' : 'none';
            }}
        }}

        function applyFilters() {{
            const interestFilter = document.querySelector('.filter-btn.active')?.id || 'filter-all';
            const categoryFilter = document.getElementById('category-filter').value;
            const rows = document.querySelectorAll('#dataTable tbody tr');
            let visibleCount = 0;
            rows.forEach(row => {{
                const interest = row.getAttribute('data-interest') === 'true';
                const category = row.getAttribute('data-category');
                let show = true;
                if (interestFilter === 'filter-interesting' && !interest) show = false;
                if (interestFilter === 'filter-not-interesting' && interest) show = false;
                if (categoryFilter !== 'all' && category !== categoryFilter) show = false;
                row.style.display = show ? '' : 'none';
                if (show) visibleCount++;
            }});
            document.getElementById('record-count').innerText = visibleCount;
        }}

        function updateActiveFiltersDisplay() {{
            let interestText = '';
            const interestFilter = document.querySelector('.filter-btn.active')?.id || 'filter-all';
            if (interestFilter === 'filter-interesting') interestText = 'Только интересные';
            else if (interestFilter === 'filter-not-interesting') interestText = 'Только неинтересные';
            else interestText = 'Все';
            
            const categorySelect = document.getElementById('category-filter');
            let categoryText = categorySelect.options[categorySelect.selectedIndex]?.text || 'Все категории';
            
            let filtersApplied = [];
            if (interestFilter !== 'filter-all') filtersApplied.push(`Интерес: ${{interestText}}`);
            if (categorySelect.value !== 'all') filtersApplied.push(`Категория: ${{categoryText}}`);
            
            const infoDiv = document.getElementById('active-filters-info');
            if (filtersApplied.length === 0) {{
                infoDiv.innerHTML = '✅ Фильтры не применены. Показаны все записи.';
            }} else {{
                infoDiv.innerHTML = `🔍 Активные фильтры: ${{filtersApplied.join(' | ')}}`;
            }}
        }}

        function downloadExcel() {{
            let interestFilter = 'all';
            if (document.getElementById('filter-interesting').classList.contains('active')) interestFilter = 'interesting';
            if (document.getElementById('filter-not-interesting').classList.contains('active')) interestFilter = 'not-interesting';
            const categoryFilter = document.getElementById('category-filter').value;

            const filteredRecords = allRecords.filter(rec => {{
                const isInteresting = rec.is_interesting === true;
                if (interestFilter === 'interesting' && !isInteresting) return false;
                if (interestFilter === 'not-interesting' && isInteresting) return false;
                const category = rec.predicted_category || 'Не классифицировано';
                if (categoryFilter !== 'all' && category !== categoryFilter) return false;
                return true;
            }});

            const isMainMode = document.getElementById('show-main-columns').classList.contains('active');
            let columnIndicesToExport;
            if (isMainMode) {{
                columnIndicesToExport = [0, ...mainColumnIndices];
            }} else {{
                columnIndicesToExport = Array.from({{length: headers.length + 1}}, (_, i) => i);
            }}

            const sheetData = [];
            const headerRow = ['№'];
            for (let i = 0; i < headers.length; i++) {{
                if (columnIndicesToExport.includes(i+1)) {{
                    headerRow.push(headers[i]);
                }}
            }}
            sheetData.push(headerRow);

            filteredRecords.forEach((rec, idx) => {{
                const row = [];
                for (let col = 0; col <= headers.length; col++) {{
                    if (!columnIndicesToExport.includes(col)) continue;
                    if (col === 0) {{
                        row.push(idx + 1);
                        continue;
                    }}
                    const fieldKey = fieldKeys[col-1];
                    let value = rec[fieldKey];
                    if (value === undefined || value === null) value = '';

                    if (fieldKey === 'number') {{
                        const url = rec.url || '';
                        const displayValue = String(value);
                        if (url) {{
                            // Создаём гиперссылку с синим цветом и подчёркиванием
                            row.push({{
                                v: displayValue,
                                l: {{ Target: url }},
                                s: {{
                                    font: {{
                                        color: {{ argb: "FF0000FF" }},
                                        underline: true
                                    }}
                                }}
                            }});
                        }} else {{
                            row.push(displayValue);
                        }}
                    }}
                    else if (fieldKey === 'is_interesting') {{
                        row.push(value ? "Да" : "Нет");
                    }}
                    else if (fieldKey === 'positions' && Array.isArray(value)) {{
                        const lines = value.map(pos => {{
                            const name = pos.name || '';
                            const price = pos.price ? ` - ${{pos.price}}` : '';
                            const quantity = pos.quantity ? ` - ${{pos.quantity}}` : '';
                            const unit = pos.unit ? ` ${{pos.unit}}` : '';
                            return `${{name}}${{price}}${{quantity}}${{unit}}`;
                        }});
                        row.push(lines.join('\\n'));
                    }}
                    else if (typeof value === 'object') {{
                        row.push(JSON.stringify(value, null, 2));
                    }}
                    else {{
                        row.push(String(value));
                    }}
                }}
                sheetData.push(row);
            }});

            // Строим лист с поддержкой стилей (xlsx-js-style)
            const wsData = [];
            const hyperlinkCells = [];

            sheetData.forEach((rowArr, rIdx) => {{
                const wsRow = [];
                rowArr.forEach((cell, cIdx) => {{
                    if (cell && typeof cell === 'object' && 'v' in cell) {{
                        wsRow.push(cell.v);
                        hyperlinkCells.push({{ row: rIdx, col: cIdx, url: cell.l.Target }});
                    }} else {{
                        wsRow.push(cell);
                    }}
                }});
                wsData.push(wsRow);
            }});

            const ws = XLSX.utils.aoa_to_sheet(wsData);

            // Применяем синий цвет + подчёркивание к ячейкам с номером тендера
            hyperlinkCells.forEach(({{ row, col, url }}) => {{
                const cellRef = XLSX.utils.encode_cell({{ r: row, c: col }});
                if (ws[cellRef]) {{
                    ws[cellRef].l = {{ Target: url }};
                    ws[cellRef].s = {{
                        font: {{
                            color: {{ rgb: "0000FF" }},
                            underline: true
                        }}
                    }};
                }}
            }});

            const wb = XLSX.utils.book_new();
            XLSX.utils.book_append_sheet(wb, ws, "Тендеры");
            XLSX.writeFile(wb, `tenders_filtered_${{new Date().toISOString().slice(0,19).replace(/:/g, '-')}}.xlsx`);
        }}

        document.getElementById('filter-all').addEventListener('click', function() {{
            document.querySelectorAll('.filter-btn').forEach(btn => btn.classList.remove('active'));
            this.classList.add('active');
            applyFilters();
            updateActiveFiltersDisplay();
        }});
        document.getElementById('filter-interesting').addEventListener('click', function() {{
            document.querySelectorAll('.filter-btn').forEach(btn => btn.classList.remove('active'));
            this.classList.add('active');
            applyFilters();
            updateActiveFiltersDisplay();
        }});
        document.getElementById('filter-not-interesting').addEventListener('click', function() {{
            document.querySelectorAll('.filter-btn').forEach(btn => btn.classList.remove('active'));
            this.classList.add('active');
            applyFilters();
            updateActiveFiltersDisplay();
        }});
        document.getElementById('category-filter').addEventListener('change', function() {{
            applyFilters();
            updateActiveFiltersDisplay();
        }});
        document.getElementById('reset-filters').addEventListener('click', function() {{
            document.getElementById('filter-all').click();
            document.getElementById('category-filter').value = 'all';
            applyFilters();
            updateActiveFiltersDisplay();
        }});
        document.getElementById('export-excel').addEventListener('click', downloadExcel);
        document.getElementById('show-all-columns').addEventListener('click', function() {{
            document.getElementById('show-all-columns').classList.add('active');
            document.getElementById('show-main-columns').classList.remove('active');
            showAllColumns();
        }});
        document.getElementById('show-main-columns').addEventListener('click', function() {{
            document.getElementById('show-main-columns').classList.add('active');
            document.getElementById('show-all-columns').classList.remove('active');
            showMainColumns();
        }});

        showAllColumns();
        applyFilters();
        updateActiveFiltersDisplay();
    </script>
    '''

    # Итоговый HTML
    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Сводная таблица тендеров</title>
    <style>
        html, body {{
            height: 100%;
            margin: 0;
            padding: 0;
        }}
        body {{
            display: flex;
            flex-direction: column;
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background-color: #f5f5f5;
        }}
        h1 {{
            color: #333;
            text-align: center;
            margin: 20px 20px 0 20px;
        }}
        .filters {{
            background: white;
            padding: 15px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            margin: 20px;
            flex-shrink: 0;
        }}
        .table-container {{
            flex: 1;
            overflow-x: auto;
            margin: 0 20px 20px 20px;
            border-radius: 8px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
            background-color: white;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 14px;
            min-width: 1200px;
        }}
        th {{
            background-color: #2c3e50;
            color: white;
            padding: 14px 12px;
            text-align: left;
            cursor: pointer;
            user-select: none;
            font-weight: 600;
            white-space: nowrap;
        }}
        th:hover {{
            background-color: #34495e;
        }}
        th::after {{
            content: ' ↕️';
            font-size: 0.8em;
            opacity: 0.7;
            margin-left: 5px;
        }}
        td {{
            padding: 12px;
            border-bottom: 1px solid #e0e0e0;
            vertical-align: top;
            line-height: 1.4;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        tr:hover {{
            background-color: #f1f4f8;
        }}
        a {{
            color: #2980b9;
            text-decoration: none;
        }}
        a:hover {{
            text-decoration: underline;
            color: #1a5276;
        }}
        a:visited {{
            color: #7d3c98;
        }}
        .status-active {{
            color: #27ae60;
            font-weight: bold;
        }}
        .status-completed {{
            color: #7f8c8d;
        }}
        .status-cancelled {{
            color: #e74c3c;
        }}
        .date {{
            font-weight: 500;
            white-space: nowrap;
        }}
        .footer {{
            flex-shrink: 0;
            padding: 15px;
            text-align: right;
            color: #555;
            font-size: 0.9em;
            background-color: white;
            margin: 0 20px 20px 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .filter-btn {{
            background-color: #e0e0e0;
            border: none;
            padding: 6px 12px;
            margin: 0 5px;
            border-radius: 4px;
            cursor: pointer;
            font-size: 14px;
            transition: all 0.2s;
        }}
        .filter-btn.active {{
            background-color: #2c3e50;
            color: white;
        }}
        .filter-btn:hover {{
            background-color: #bdc3c7;
        }}
        select {{
            padding: 6px 12px;
            border-radius: 4px;
            border: 1px solid #ccc;
            font-size: 14px;
        }}
    </style>
</head>
<body>
    <h1>📊 Сводная информация по тендерам</h1>
    {filter_buttons_html}
    <div class="table-container">
        <table id="dataTable">
            <thead>
                <tr>{headers_html}<tr>
            </thead>
            <tbody>
                {''.join(rows_html)}
            </tbody>
        </table>
    </div>
    <div class="footer">
        <span>📅 Сгенерировано {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')} | </span>
        <span>📋 Всего записей: <strong>{len(records)}</strong></span>
    </div>
    {js_filter}
</body>
</html>"""
    return html


if __name__ == "__main__":
    import sys
    import webbrowser

    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    output_html = os.path.join(EXTRACTED_TEXT_FOLDER, "tenders_summary.html")
    generate_html_table(EXTRACTED_TEXT_FOLDER, output_html, jsonl_filename="all_tenders_data_20260331_020426.jsonl")
    
    output_excel = os.path.join(EXTRACTED_TEXT_FOLDER, "tenders_summary.xlsx")
    generate_excel_table(EXTRACTED_TEXT_FOLDER, output_excel, jsonl_filename="all_tenders_data_20260331_020426.jsonl")

    print(f"HTML таблица сгенерирована: {output_html}")
    print(f"Excel таблица сгенерирована: {output_excel}")
    webbrowser.open(f"file://{output_html}")